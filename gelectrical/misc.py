#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
# bill_dialog.py
#  
#  Copyright 2014 Manu Varkey <manuvarkey@gmail.com>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

import subprocess, threading, os, posixpath, platform, logging, math, cairo, copy, time, pathlib
import base64
from collections.abc import MutableMapping
from uuid import uuid4 as uuid
from urllib.parse import urlparse
from urllib.request import url2pathname
import pandas as pd
import numpy as np
from scipy.interpolate import interp1d

from gi.repository import Gtk, Gdk, GLib, Pango, PangoCairo, GdkPixbuf
import openpyxl

# Setup logger object
log = logging.getLogger(__name__)


## GLOBAL CONSTANTS

# Program name
PROGRAM_NAME = 'GElectrical'
PROGRAM_AUTHOR = 'kavilgroup'
PROGRAM_VER = '1'
APPID = "com.kavilgroup.gelectrical"
# Paths updated from __init__
USER_LIBRARY_DIR = ''
# Error codes used for displaying info in main window
ERROR = -1
WARNING = -2
OK = 0
INFO = 0
# Defalt colors for models
COLOR_NORMAL = '#000000'
COLOR_INACTIVE = '#888a85'
COLOR_OVERLAY_BG = '#ad7fa880'
COLOR_OVERLAY_TEXT = '#00000080'
COLOR_GRID = '#00ff00FF'
COLOR_GRID_MINOR = '#00ff0020'
COLOR_SELECTED = '#729fcf'
COLOR_SELECTED_WARNING = '#cc0000'
COLOR_SELECTION_BAND = '#729fcf'
COLOR_PLOT_BG = '#FFFFFF'

def set_dark_mode_drawing_values():
    global COLOR_NORMAL, COLOR_OVERLAY_TEXT, COLOR_PLOT_BG
    COLOR_NORMAL = '#FFFFFF'
    COLOR_OVERLAY_TEXT = '#FFFFFF80'
    COLOR_PLOT_BG = '#000000'

def reset_dark_mode_drawing_values():
    global COLOR_NORMAL, COLOR_OVERLAY_TEXT, COLOR_PLOT_BG
    COLOR_NORMAL = '#000000'
    COLOR_OVERLAY_TEXT = '#00000080'
    COLOR_PLOT_BG = '#FFFFFF'

def check_dark_mode():
    global COLOR_NORMAL
    if COLOR_NORMAL == '#FFFFFF':
        return True
    else:
        return False
    
# Default stroke widths
STROKE_WIDTH_NORMAL = 0.5/0.3527
STROKE_WIDTH_THIN = 0.25/0.3527
STROKE_WIDTH_THICK = 0.7/0.3527
STROKE_WIDTH_EXTRATHICK = 1.4/0.3527
STROKE_WIDTH_GRID = 1
STROKE_WIDTH_SELECTED = 3
STROKE_WIDTH_SELECTION_BAND = 2
RECTANGLE_RADIUS = 5
# Timeout for killing Latex subprocess
PROCESS_TIMEOUT = 300 # 5 minutes
# Timeout of auto hiding message
MESSAGE_TIMEOUT = 10
# String used for checking file version
PROJECT_FILE_VER = 'GELECTRICAL_FILE_REFERENCE_VER_0'
# Default settings
PAGE_WIDTH = 420  # for A3
PAGE_HEIGHT = 297  # for A3
GRID_WIDTH = 16
SELECT_PORT_RECT = 16
SCHEM_FONT_FACE = 'osifont'
SCHEM_FONT_SIZE = 7  # Keep minimum of 7 point (x 0.3527 in mm)
SCHEM_FONT_WEIGHT = Pango.Weight.MEDIUM
SCHEM_FONT_WEIGHT_BOLD = Pango.Weight.HEAVY
TITLE_FONT_SIZE_SMALL = 6
TITLE_FONT_SIZE = 8
SCHEM_FONT_SPACING = 10
SELECT_DRAW_PATTERN = (30,10)
WIRE_ADD_DRAW_PATTERN = (5,5)
GROUP_DRAW_PATTERN = (20,5,5,5)
FIELD_CAPTION_WIDTH = 100
FIELD_DIALOG_CAPTION_WIDTH = 300
FIELD_UNIT_WIDTH = 25
WINDOW_WIDTH = 1600
WINDOW_HEIGHT = 900
# Status codes for drawing view
MODE_DEFAULT = 0
MODE_SELECTION = 1
MODE_INSERT = 2
MODE_ADD_WIRE = 3
# Graph codes
## graph types
GRAPH_DATATYPE_FREE = 0
GRAPH_DATATYPE_PROFILE = 1
GRAPH_DATATYPE_POLYGON= 2
GRAPH_DATATYPE_MARKER = 3
## graph settings
GRAPH_FONT_FACE = 'monospace'
GRAPH_FONT_SIZE = 10
GRAPH_LOAD_TIME_LIMITS = (0,23,1)
GRAPH_LOAD_CURRENT_LIMITS = (0,1.5,0.05)
GRAPH_PROT_TIME_LIMITS = (0.01,3600,60,'log')
GRAPH_PROT_CURRENT_LIMITS = (1,100e3,100,'log')
GRAPH_PROT_CURRENT_LIMITS_G = (0.01,10e3,100,'log')
GRAPH_COLORS = ('#ef2929', '#F7BD03', '#729fcf', '#a40000', '#CF9033', '#204a87', '#4e9a06', '#ce5c00')
# Report parameters
REPORT_FONT_FACE = 'monospace'
REPORT_FONT_SIZE = 10
REPORT_GRAPH_FONT_FACE = 'monospace'
REPORT_GRAPH_FONT_SIZE = 10
# Power model types
POWER_MODEL_POWERFLOW = 0
POWER_MODEL_LINEFAULT = 1
POWER_MODEL_GROUNDFAULT = 2
# Classification of elements
ADVANCED_ELEMENTS = ('element_asymmetric_load', 'element_ward', 'element_xward',
                     'element_transformer3w', 'element_impedance', 'element_shunt')
REFERENCE_CODES = ('element_reference', 'element_reference_box')
LOADPROFILE_CODES = ('element_load', 'element_asymmetric_load', 'element_single_phase_load', 
                     'element_staticgenerator', 'element_single_phase_staticgenerator', 'element_async_motor',
                     'element_async_motor_3ph', 'element_async_motor_1ph')
LINE_ELEMENT_CODES = ('element_line', 'element_line_cable', 'element_line_custom', 'element_line_bus')
TRAFO_ELEMENT_CODES = ('element_transformer', 'element_transformer3w')
MOTOR_ELEMENT_CODES = ('element_async_motor_3ph', 'element_async_motor_1ph')
LOAD_ELEMENT_CODES = ('element_load', 'element_asymmetric_load', 'element_single_phase_load', 'element_async_motor_3ph', 'element_async_motor_1ph')
SWITCH_ELEMENT_CODES = ('element_switch', 'element_circuitbreaker', 'element_fuse', 'element_contactor')
PROTECTION_ELEMENT_CODES = ('element_circuitbreaker', 'element_fuse', 'element_contactor')
DAMAGE_ELEMENT_CODES = ('element_async_motor_3ph', 'element_async_motor_1ph','element_line','element_line_cable','element_line_custom', 'element_line_bus','element_transformer')
SUPPLY_ELEMENT_CODES = ('element_grid', 'element_generator', 'element_staticgenerator', 'element_single_phase_staticgenerator')
DISPLAY_ELEMENT_CODES = ('element_display_node', )
NON_ELEMENT_CODES  = (*REFERENCE_CODES, *DISPLAY_ELEMENT_CODES, 'element_display_text', 'element_wire', 'element_assembly')
# Defaults
DEFAULT_LOAD_PROFILE = {'load_prof_1': ['Flat', [{'mode':GRAPH_DATATYPE_PROFILE, 'title':'Default', 'xval':[0,23], 'yval':[1,1]}]],}
# Constants
PI = 3.141593
M = 8  # Multiplier to be used in drawings
POINT_TO_MM = 0.3527
# Paper sizes in mm, (width, height, (width fold marks), (height fold marks))
paper_sizes = {'A4 Landscape': (297, 210),
               'A4 Portrait': (210, 297),
               'A3': (420, 297),
               'A2': (594, 420),
               'A1': (841, 594),
               'A0': (1189, 841), }

               
## UNDO MANAGEMENT
    
class _Action:
    ''' This represents an action which can be done and undone.
    
    It is the result of a call on an undoable function and has
    three methods: ``do()``, ``undo()`` and ``text()``.  The first value
    returned by the internal call in ``do()`` is the value which will 
    subsequently be returned by ``text``.  Any remaining values are 
    returned by ``do()``.
    '''

    def __init__(self, generator, args, kwargs):
        self._generator = generator
        self.args = args
        self.kwargs = kwargs
        self._text = ''

    def do(self):
        'Do or redo the action'
        self._runner = self._generator(*self.args, **self.kwargs)
        rets = next(self._runner)
        if isinstance(rets, tuple):
            self._text = rets[0]
            return rets[1:]
        elif rets is None:
            self._text = ''
            return None
        else:
            self._text = rets
            return None

    def undo(self):
        'Undo the action'
        try:
            next(self._runner)
        except StopIteration:
            pass
        # Delete it so that its not accidentally called again
        del self._runner

    def text(self):
        'Return the descriptive text of the action'
        return self._text

def undoable(generator):
    ''' Decorator which creates a new undoable action type. 
    
    This decorator should be used on a generator of the following format::
    
        @undoable
        def operation(*args):
            do_operation_code
            yield 'descriptive text'
            undo_operator_code
    '''

    def inner(*args, **kwargs):
        action = _Action(generator, args, kwargs)
        ret = action.do()
        args[0].stack.append(action)
        if isinstance(ret, tuple):
            if len(ret) == 1:
                return ret[0]
            elif len(ret) == 0:
                return None
        return ret

    return inner
    
class _Group:
    ''' A undoable group context manager. '''

    def __init__(self, desc, stack):
        self._desc = desc
        self._stack = []
        self.stack = stack

    def __enter__(self):
        self.stack.setreceiver(self._stack)

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self.stack.resetreceiver()
            self.stack.append(self)
        return False

    def undo(self):
        for undoable in reversed(self._stack):
            undoable.undo()

    def do(self):
        for undoable in self._stack:
            undoable.do()

    def text(self):
        return self._desc.format(count=len(self._stack))

def group(self, desc):
    ''' Return a context manager for grouping undoable actions. 
    
    All actions which occur within the group will be undone by a single call
    of `stack.undo
    '''
    return _Group(desc, self.stack)

def get_undoable_set_field(stack, refresh, element):
    # Special class implementing undo for set text function
    class UndoableSetTextValue:
        def __init__(self, stack, refresh):
            self.stack = stack
            self.refresh = refresh
    
        @undoable
        def set_text_field_value_undo(self, code, value):
            oldval = element.get_text_field(code)['value']
            modified = element.set_text_field_value(code, value)
            if self.refresh:
                self.refresh() 
            yield "Modify '{}' field  of '{}' element".format(code, element.name)
            # Undo action
            element.set_text_field_value(code, oldval)
            if modified:
                for code, value in modified.items():
                    element.set_text_field_value(code, value)
            if self.refresh:
                self.refresh() 
    return UndoableSetTextValue(stack, refresh).set_text_field_value_undo


def get_undoable_set_field_multiple(stack, refresh, elements, multiselect_fields):
    # Special class implementing undo for set text function
    class UndoableSetTextValue:
        def __init__(self, stack, refresh):
            self.stack = stack
            self.refresh = refresh
    
        @undoable
        def set_text_field_value_undo(self, code, value):
            oldvals = []
            for element in elements:
                oldval = element.get_text_field(code)['value']
                element.set_text_field_value(code, value)
                oldvals.append(oldval)
            multiselect_fields[code]['value'] = value
            if self.refresh:
                self.refresh() 
            yield "Modify '{}' field  of '{}' (multiple) elements".format(code, elements[0].name)
            # Undo action
            for oldval, element in zip(oldvals, elements):
                element.set_text_field_value(code, oldval)
            if self.refresh:
                self.refresh() 
    return UndoableSetTextValue(stack, refresh).set_text_field_value_undo
    
## GLOBAL VARIABLES


## GLOBAL CLASSES

            
class SpreadsheetDialog:
    """Dialog for manage input and output of spreadsheets"""
   
    def __init__(self, parent, filename, columntypes, captions, dimensions = None, allow_formula=False):
        """Initialise SpreadsheetDialog class
        
            Arguments:
                parent: Parent widget (Main window)
                filename: 
                columntypes: Data types of columns. 
                             Takes following values:
                                misc.MEAS_NO: Integer
                                misc.MEAS_L: Float
                                misc.MEAS_DESC: String
                                misc.MEAS_CUST: Value omited
                dimensions: List of two lists passing column widths and expand properties
                allow_formula: Reads all values as string
        """
        log.info('SpreadsheetDialog - Initialise')
        # Setup variables
        self.parent = parent
        self.filename = filename
        self.captions = captions
        self.columntypes = columntypes
        self.dimensions = dimensions
        self.allow_formula = allow_formula
        
        self.top = 0
        self.bottom = 0
        self.left = 0
        self.right = 0
        self.values = []
        self.spreadsheet = None
        self.sheet = ''

        # Setup dialog window
        self.builder = Gtk.Builder()
        self.builder.add_from_file(abs_path("interface","spreadsheetdialog.glade"))
        self.window = self.builder.get_object("dialog")
        self.window.set_transient_for(self.parent)
        self.window.set_default_size(1100,600)
        self.builder.connect_signals(self)

        # Get required objects
        self.combo = self.builder.get_object("combobox_sheet")
        self.combo_store = self.builder.get_object("liststore_combo")
        self.tree = self.builder.get_object("treeview_schedule")
        self.entry_top = self.builder.get_object("entry_top")
        self.entry_bottom = self.builder.get_object("entry_bottom")
        self.entry_left = self.builder.get_object("entry_left")
        self.entry_right = self.builder.get_object("entry_right")
        
        # Setup treeview
        self.columns = []
        self.cells = []
        # Setup row number column
        cell_row = Gtk.CellRendererText()
        column_row = Gtk.TreeViewColumn('', cell_row)
        column_row.add_attribute(cell_row, "markup", 0)
        column_row.set_min_width(50)
        column_row.set_fixed_width(50)
        cell_row.props.wrap_width = 50
        cell_row.props.background = COLOR_INACTIVE
        self.cells.append(cell_row)
        self.columns.append(column_row)
        self.tree.append_column(column_row)
        # Setup remaining columns
        for c_no, caption  in enumerate(self.captions,1):
            cell = Gtk.CellRendererText()
            column = Gtk.TreeViewColumn(caption, cell)
            column.props.sizing = Gtk.TreeViewColumnSizing.AUTOSIZE
            column.connect("notify", self.on_wrap_column_resized, cell)
            column.add_attribute(cell, "text", c_no)
            self.cells.append(cell)
            self.columns.append(column)
            self.tree.append_column(column)
        # Setup dimensions
        if dimensions is not None:
            self.setup_column_props(*dimensions)
        # Setup liststore model
        types = [str] + [str]*len(self.columntypes)
        self.store = Gtk.ListStore(*types)
        self.tree.set_model(self.store)
        # Misc options
        self.tree.set_grid_lines(3)
        self.tree.set_enable_search(True)
        search_cols = [no for no,x in enumerate(self.columntypes,1) if x == str]
        self.tree.set_search_equal_func(self.equal_func, [0,1,2,3])
        
        # Read file into spreadsheet object
        if filename is not None:
            try:
                self.spreadsheet = Spreadsheet(filename)
            except:
                self.spreadsheet = None
                log.warning('SpreadsheetDialog - Spreadsheet could not be read - ' + filename)
                
        # Setup combobox
        if self.spreadsheet:
            sheets = self.spreadsheet.sheets()
            for sheet in sheets:
                self.combo_store.append([sheet])
            if sheets:
                self.combo.set_active_id(sheets[0])
                self.update()
            

    def run(self):
        """Display dialog box and return data model
        
            Returns:
                Data Model on Ok
                [] on Cancel
        """
        self.window.show_all()
        response = self.window.run()
        self.window.destroy()

        if response == 1 and self.spreadsheet:
            log.info('SpreadsheetDialog - run - Response Ok')
            return self.values
        else:
            log.info('SpreadsheetDialog - run - Response Cancel')
            return []
    
    def update(self):
        """Update contents from input values"""
        log.info('SpreadsheetDialog - Update')
        
        # Read if sheet changed
        sheet = self.combo_store[self.combo.get_active_iter()][0]
        if sheet != self.sheet:
            self.sheet = sheet
            self.spreadsheet.set_active_sheet(self.sheet)
            self.entry_top.set_text('1')
            self.entry_bottom.set_text(str(self.spreadsheet.length()+1))
            self.entry_left.set_text('1')
            
        # Read values of entries
        self.top = int(self.entry_top.get_text())
        self.bottom = int(self.entry_bottom.get_text())
        self.left = int(self.entry_left.get_text())
        
        # Set values
        self.entry_right.set_text(str(self.left + len(self.columntypes)))
        
        # Read spreadsheet
        self.values = self.spreadsheet.read_rows(self.columntypes, start=self.top-1, end=self.bottom-1, left=self.left-1, allow_formula=self.allow_formula)
        
        # Update store
        self.store.clear()
        for slno, value in enumerate(self.values, self.top):
            formated_value = [str(x) if x != 0 else '' for x in value]
            self.store.append(['<b>' + str(slno) + '</b>'] + formated_value)
                
    def setup_column_props(self, widths, expandables):
        """Set column properties
            Arguments:
                widths: List of column widths type-> [int, ...]. None values are skiped.
                expandables: List of expand property type-> [bool, ...]. None values are skiped.
        """
        for column, cell, width, expandable in zip(self.columns[1:], self.cells[1:], widths, expandables):
            if width != None:
                column.set_min_width(width)
                column.set_fixed_width(width)
                cell.props.wrap_width = width
            if expandable != None:
                column.set_expand(expandable)
    
    def equal_func(self, model, column, key, iter, cols):
        """Equal function for interactive search"""
        search_string = ''
        for col in cols:
            search_string += ' ' + model[iter][col].lower()
        for word in key.split():
            if word.lower() not in search_string:
                return True
        return False
        
    def on_wrap_column_resized(self, column, pspec, cell):
        """ Automatically adjust wrapwidth to column width"""

        width = column.get_width() - 5
        oldwidth = cell.props.wrap_width
        
        if width > 0 and width != oldwidth:
            cell.props.wrap_width = width
            # Force redraw of treeview
            GLib.idle_add(column.queue_resize)
    
    # Callbacks
    
    def onRefreshClicked(self, button):
        """Refresh screen on button click"""
        
        # Sanitise entries
        if self.entry_top.get_text() == '':
            self.entry_top.set_text('1')
        if self.entry_bottom.get_text() == '':
            self.entry_bottom.set_text('1')
        if self.entry_left.get_text() == '':
            self.entry_left.set_text('1')
        
        if self.spreadsheet:
            self.update()
    
    def onEntryEditedNum(self, entry):
        """Treeview cell renderer for editable number field
        
            User Data:
                column: column in ListStore being edited
        """
        new_text = entry.get_text()
        num = ''
        if new_text != '':
            try:  # check whether item evaluates fine
                num = int(new_text)
                if num <= 0:
                   num = 1
            except:
                log.warning("SpreadsheetDialog - onEntryEditedNum - evaluation of [" 
                + new_text + "] failed")
        entry.set_text(str(num))

            
class Spreadsheet:
    """Manage input and output of spreadsheets"""
    
    def __init__(self, filename=None):
        if filename is not None:
            self.spreadsheet = openpyxl.load_workbook(filename)
        else:
            self.spreadsheet = openpyxl.Workbook()
        self.sheet = self.spreadsheet.active
    
    def save(self, filename):
        """Save worksheet to file"""
        self.spreadsheet.save(filename)
        
    # Sheet management
    
    def new_sheet(self):
        """Create a new sheet to spreadsheet and set as active"""
        self.sheet = self.spreadsheet.create_sheet()  
            
    def sheets(self):
        """Returns a list of sheetnames"""
        return self.spreadsheet.sheetnames
        
    def length(self):
        """Get number of rows in sheet"""
        return self.sheet.max_row
        
    def set_title(self, title):
        """Set title of sheet"""
        self.sheet.title = title
        
    def set_page_settings(self, orientation='portrait', papersize='A4', font=None, print_title_rows = None):
        # Orientation
        if orientation == 'portrait':
            self.sheet.page_setup.orientation = openpyxl.worksheet.worksheet.Worksheet.ORIENTATION_PORTRAIT
        elif orientation == 'landscape':
            self.sheet.page_setup.orientation = openpyxl.worksheet.worksheet.Worksheet.ORIENTATION_LANDSCAPE
            
        # Paper size
        if papersize == 'A4':
            self.sheet.page_setup.paperSize = openpyxl.worksheet.worksheet.Worksheet.PAPERSIZE_A4
            
        # Print title rows
        self.sheet.print_title_rows = print_title_rows
            
        # General settings
        self.sheet.page_setup.fitToPage = True
        self.sheet.page_setup.fitToHeight = 99
        self.sheet.page_setup.fitToWidth = 1
        self.sheet.print_options.horizontalCentered = True

        # Set font
        if font:
            for row in range(1, self.length()+1):
                for cell in self.sheet[row]:
                    font_style = copy.copy(cell.font)
                    font_style.name = font
                    cell.font = font_style

    def set_column_widths(self, widths):
        """Set column widths of sheet"""
        for column, width in enumerate(widths, 1):
            col_letter = openpyxl.utils.get_column_letter(column)
            self.sheet.column_dimensions[col_letter].width = width
        
    def set_active_sheet(self, sheetref):
        """Set active sheet of spreadsheet"""
        sheetname = ''
        sheetno = None
        if type(sheetref) is int:
            sheetno = sheetref
        elif type(sheetref) is str:
            sheetname = sheetref
        
        if sheetname in self.sheets():
            self.sheet = self.spreadsheet[sheetname]
        elif sheetno is not None and sheetno < len(self.sheets()):
            self.sheet = self.spreadsheet[self.sheets()[sheetno]]
    
    def set_style(self, row, col, bold=False, wrap_text=True, horizontal='general', vertical='bottom', fill=None):
        """Set style of individual cell"""
        font = openpyxl.styles.Font(bold=bold)
        alignment = openpyxl.styles.Alignment(wrap_text=wrap_text, horizontal=horizontal, vertical=vertical)

        self.sheet.cell(row=row, column=col).font = font
        self.sheet.cell(row=row, column=col).alignment = alignment
        
        if fill == '#FFFFFF':
            patternfill = openpyxl.styles.PatternFill()
            self.sheet.cell(row=row, column=col).fill = patternfill
        elif fill is not None:
            patternfill = openpyxl.styles.PatternFill(start_color=fill[1:], end_color=fill[1:], fill_type='solid')
            self.sheet.cell(row=row, column=col).fill = patternfill
        
    # Data addition functions
            
    def append(self, ss_obj):
        """Append an sheet to current sheet"""
        sheet = ss_obj.spreadsheet.active
        rowcount = self.length()
        for row_no, row in enumerate(sheet.rows, 1):
            for col_no, cell in enumerate(row, 1):
                self.sheet.cell(row=row_no+rowcount, column=col_no).value = cell.value
                if cell.has_style:
                    self.sheet.cell(row=row_no+rowcount, column=col_no).font = copy.copy(cell.font)
                    self.sheet.cell(row=row_no+rowcount, column=col_no).border = copy.copy(cell.border)
                    self.sheet.cell(row=row_no+rowcount, column=col_no).fill = copy.copy(cell.fill)
                    self.sheet.cell(row=row_no+rowcount, column=col_no).number_format = copy.copy(cell.number_format)
                    self.sheet.cell(row=row_no+rowcount, column=col_no).protection = copy.copy(cell.protection)
                    self.sheet.cell(row=row_no+rowcount, column=col_no).alignment = copy.copy(cell.alignment)

    def append_data(self, data, bold=False, italic=False, wrap_text=True, horizontal='general', vertical='bottom', fill = None):
        """Append data to current sheet"""
        rowcount = self.length()
        self.insert_data(data, rowcount+1, 1, bold, italic,wrap_text, horizontal, vertical, fill)
    
    def insert_data(self, data, start_row=1, start_col=1, bold=False, italic=False, wrap_text=True, horizontal='general', vertical='bottom', fill=None):
        """Insert data to current sheet"""
        # Setup styles
        font = openpyxl.styles.Font(bold=bold, italic=italic)
        
        if fill == '#FFFFFF':
            patternfill = openpyxl.styles.PatternFill()
        elif fill is not None:
            patternfill = openpyxl.styles.PatternFill(start_color=fill[1:], end_color=fill[1:], fill_type='solid')
            
        alignment = openpyxl.styles.Alignment(wrap_text=wrap_text, horizontal=horizontal, vertical=vertical)
        # Apply data and styles
        for row_no, row in enumerate(data, start_row):
            for col_no, value in enumerate(row, start_col):
                self.sheet.cell(row=row_no, column=col_no).value = value
                self.sheet.cell(row=row_no, column=col_no).font = font
                
                if fill:
                    self.sheet.cell(row=row_no, column=col_no).fill = patternfill
                    
                self.sheet.cell(row=row_no, column=col_no).alignment = alignment
                
    def add_merged_cell(self, value, row=None, width=2, bold=False, wrap_text=True, horizontal='center', start_column=1):
        """Add a merged cell of prescrbed width"""
        if row is None:
            rowstart = self.length() + 1
        else:
            rowstart = row
        self.sheet.merge_cells(start_row=rowstart,start_column=start_column,end_row=rowstart,end_column=start_column+width-1)
        self.__setitem__([rowstart,start_column], value)
        self.set_style(rowstart, start_column, bold, wrap_text, horizontal)
    
    def __setitem__(self, index, value):
        """Set an individual cell"""
        self.sheet.cell(row=index[0], column=index[1]).value = value
        
    def __getitem__(self, index):
        """Set an individual cell"""
        return self.sheet.cell(row=index[0], column=index[1]).value
            
    # Bulk read functions
    
    def read_rows(self, columntypes = [], start=0, end=-1, left=0, allow_formula=False):
        """Read and validate selected rows from current sheet"""
        # Get count of rows
        rowcount = self.length()
        if end < 0 or end >= rowcount:
            count_actual = rowcount
        else:
            count_actual = end
        
        items = []
        for row in range(start, count_actual):
            cells = []
            skip = 0  # No of columns to be skiped ex. breakup, total etc...
            for columntype, i in zip(columntypes, list(range(left, len(columntypes)+left))):
                cell = self.sheet.cell(row = row + 1, column = i - skip + 1).value
                if columntype == str:
                    if cell is None:
                        cell_formated = ""
                    else:
                        cell_formated = str(cell)
                elif columntype == float:
                    if cell is None:
                        if allow_formula:
                            cell_formated = '0'
                        else:
                            cell_formated = 0
                    else:
                        if allow_formula:
                            try:  # try evaluating float
                                if len(str(cell)) > 1 and str(cell)[0] == '=':
                                    formula = str(cell)[1:]
                                else:
                                    formula = str(cell)
                                evaluated = str(float(eval(formula)))
                                cell_formated = formula
                            except:
                                cell_formated = '0'
                        else:
                            try:  # try evaluating float
                                cell_formated = str(float(cell))
                            except:
                                cell_formated = 0

                elif columntype == int:
                    if cell is None:
                        if allow_formula:
                            cell_formated = '0'
                        else:
                            cell_formated = 0
                    else:
                        if allow_formula:
                            try:  # try evaluating int
                                if len(str(cell)) > 1 and str(cell)[0] == '=':
                                    formula = str(cell)[1:]
                                else:
                                    formula = str(cell)
                                evaluated = str(int(eval(formula)))
                                cell_formated = formula
                            except:
                                cell_formated = '0'
                        else:
                            try:  # try evaluating int
                                cell_formated = str(int(cell))
                            except:
                                cell_formated = 0
                else:
                    cell_formated = ''
                    log.warning("Spreadsheet - Value skipped on import - " + str((row, i)))
                if columntype is None:
                    skip = skip + 1
                cells.append(cell_formated)
            items.append(cells)
        return items

        
class ProgressRevealer:
    """Class for handling display of long running proccess"""
    
    def __init__(self, parent=None, label=None, progress=None):
        
        self.parent = parent
        self.label = label
        self.progress = progress
        # Setup data
        self.step = 0
        self.fraction = 0
        
    # Revealer functions
    
    def show(self):
        GLib.idle_add(self.parent.set_reveal_child, True)
    
    def close(self):
        GLib.idle_add(self.parent.set_reveal_child, False)
    
    # General functions
        
    def set_pulse_step(self, width):
        self.step = width
        self.fraction = 0
        
    def set_fraction(self, fraction):
        def callback():
            self.fraction = fraction
            self.progress.set_fraction(self.fraction)
            self.show()
            return False
        GLib.idle_add(callback)

    def get_fraction(self):
        return self.progress.get_fraction()
        
    def pulse(self, end=False):
        def callback():
            self.fraction += self.step
            if end:
                self.fraction = 1
            else:
                self.fraction = min(self.fraction + self.step, 1)
            self.progress.set_fraction(self.fraction)
            self.show()
            return False
        GLib.idle_add(callback)
        
    def add_message(self, message):
        GLib.idle_add(self.label.set_markup, message)
            
            
class SplashScreen:
    def __init__(self, callback, image_filename, min_splash_time=0 ):
        self.image = Gtk.Image.new_from_file(image_filename )
        self.window = Gtk.Window(Gtk.WindowType.TOPLEVEL)
        self.window.set_position(Gtk.WindowPosition.CENTER)
        self.window.set_type_hint(Gdk.WindowTypeHint.SPLASHSCREEN)
        self.window.set_gravity(Gdk.Gravity.CENTER)
        self.window.set_auto_startup_notification(False)
        self.window.set_decorated(False)
        self.window.add(self.image)
        self.min_splash_time   = time.time() + min_splash_time
        self.window.show_all()
        GLib.timeout_add_seconds(1, callback)
   
    def exit(self):
        # Make sure the minimum splash time has elapsed
        timeNow = time.time()
        if timeNow < self.min_splash_time:
            time.sleep( self.min_splash_time - timeNow )
          
        # Destroy the splash window
        self.window.destroy( )
        

class Command(object):
    """Runs a command in a seperate thread"""
    
    def __init__(self, cmd):
        """Initialises class with command to be executed"""
        self.cmd = cmd
        self.process = None

    def run(self, timeout):
        """Run set command with selected timeout"""
        def target():
            if platform.system() == 'Linux':
                self.process = subprocess.Popen(self.cmd)
            elif platform.system() == 'Windows':
                self.process = subprocess.Popen(self.cmd, shell=True)
            log.info('Sub-process spawned - ' + str(self.process.pid))
            self.process.communicate()
        thread = threading.Thread(target=target)
        thread.start()

        thread.join(timeout)
        if thread.is_alive():
            log.error('Terminating sub-process exceeding timeout - ' + str(self.process.pid))
            self.process.terminate()
            thread.join()
            return -1
        return 0

class ReferenceCounter:
    """Class for implementing reference counting"""
    
    def __init__(self, start):
        self.refs = dict()
        self.start=start
        
    def __setitem__(self, code, value):
        self.refs[code] = value
            
    def __getitem__(self, code):
        if code in self.refs:
            return self.refs[code]
        else:
            return self.start

class FieldDict(MutableMapping):
    """Convinence class to read field dictionary attributes"""

    def __init__(self, dict_var):
        self.store = dict_var

    def __getitem__(self, key):
        return self.store[key]['value']

    def __setitem__(self, key, value):
        self.store[key]['value'] = value

    def __delitem__(self, key):
        del self.store[key]

    def __iter__(self):
        return iter(self.store)
    
    def __len__(self):
        return len(self.store)
    
    __getattr__ = __getitem__

class Element:
    def __init__(self, element):
        # Copy all variables
        self.__dict__.update(vars(element))
        # Add convinence variables
        self.r = FieldDict(element.res_fields)
        self.f = FieldDict(element.fields)

## GLOBAL METHODS

def pprint(variable):
    """Pretty print for debug purposes"""
    import json
    print(json.dumps(variable, indent=2))

def display_message(parent, message, title='', message_type='info'):
    '''Gets a single user input by diplaying a dialog box
    
    Arguments:
        parent: Parent window
        message: Message to be displayed to user
        title: Dialog title text
        message_type: question/info/warning/error
    Returns:
        Returns response of dialog.
    '''
    message_type_dict = {'question': Gtk.MessageType.QUESTION,
                         'info': Gtk.MessageType.INFO,
                         'warning': Gtk.MessageType.WARNING,
                         'error': Gtk.MessageType.ERROR}
    buttons_type_dict = {'question': Gtk.ButtonsType.YES_NO,
                         'info': Gtk.ButtonsType.OK,
                         'warning': Gtk.ButtonsType.OK,
                         'error': Gtk.ButtonsType.OK}
    dialogWindow = Gtk.MessageDialog(parent,
                                     Gtk.DialogFlags.MODAL | Gtk.DialogFlags.DESTROY_WITH_PARENT,
                                     message_type_dict[message_type],
                                     buttons_type_dict[message_type],
                                     message)

    dialogWindow.set_transient_for(parent)
    dialogWindow.set_title(title)
    dialogWindow.set_default_response(Gtk.ResponseType.OK)

    dialogBox = dialogWindow.get_content_area()
    userlabel = Gtk.Label(title)
    dialogBox.pack_end(userlabel, False, False, 0)

    dialogWindow.show_all()
    response = dialogWindow.run()
    dialogWindow.destroy()
    if response == Gtk.ResponseType.OK:
        return True
    else:
        return False

def get_user_input_text(parent, message, title='', oldval=None, multiline=False):
    '''Gets a single user input by diplaying a dialog box
    
    Arguments:
        parent: Parent window
        message: Message to be displayed to user
        title: Dialog title text
        multiline: Allows multiline input is True
    Returns:
        Returns user input as a string or 'None' if user does not input text.
    '''
    dialogWindow = Gtk.MessageDialog(parent,
                                     Gtk.DialogFlags.MODAL | Gtk.DialogFlags.DESTROY_WITH_PARENT,
                                     Gtk.MessageType.QUESTION,
                                     Gtk.ButtonsType.OK_CANCEL,
                                     message)

    dialogWindow.set_transient_for(parent)
    dialogWindow.set_title(title)
    dialogWindow.set_default_response(Gtk.ResponseType.OK)

    dialogBox = dialogWindow.get_content_area()
    text = ''
    
    if multiline:
        # Function to mark first line as bold
        def mark_heading(textbuff, tag):
            start = textbuff.get_start_iter()
            end = textbuff.get_end_iter()
            textbuff.remove_all_tags(start, end)
            match = start.forward_search('\n', 0, end)
            if match != None:
                match_start, match_end = match
                textbuff.apply_tag(tag, start, match_start)

        scrolledwindow = Gtk.ScrolledWindow()
        scrolledwindow.set_hexpand(True)
        scrolledwindow.set_vexpand(True)
        scrolledwindow.set_size_request(300, 100)

        textview = Gtk.TextView()
        textbuffer = textview.get_buffer()
        dialogBox.pack_end(scrolledwindow, False, False, 0)
        scrolledwindow.add(textview)
        
        # Set old value
        if oldval != None:
            textbuffer.set_text(oldval)
        
        # Mark heading
        tag_bold = textbuffer.create_tag("bold", weight=Pango.Weight.BOLD)
        mark_heading(textbuffer, tag_bold)
        textbuffer.connect("changed", mark_heading, tag_bold)

        dialogWindow.show_all()
        response = dialogWindow.run()
        text = textbuffer.get_text(textbuffer.get_start_iter(),textbuffer.get_end_iter(), True)
    else:
        userEntry = Gtk.Entry()
        userEntry.set_activates_default(True)
        userEntry.set_size_request(50, 0)
        dialogBox.pack_end(userEntry, False, False, 0)
        
        # Set old value
        if oldval != None:
            userEntry.set_text(oldval)

        dialogWindow.show_all()
        response = dialogWindow.run()
        text = userEntry.get_text()
    dialogWindow.destroy()
    if (response == Gtk.ResponseType.OK) and (text != ''):
        return text
    else:
        return None
    
def set_widget_font(widget, fontname, fontsize=None):
    context = widget.get_pango_context() 
    font = context.get_font_description()
    font.set_family(fontname)
    if fontsize:
        font.set_size(fontsize * Pango.SCALE)
    context.set_font_description(font)
    
def file_to_uri(filename):
    if platform.system() == 'Windows':
        path = pathlib.PureWindowsPath(filename)
        uri = path.as_uri()
        # If network path
        if filename.startswith(r'\\'):
            uri = uri[0:7] + '//' + uri[7:]
    else:
        path = pathlib.Path(filename)
        uri = path.as_uri()
    return uri
    
def uri_to_file(uri):
    return url2pathname(urlparse(uri).path)
    
def abs_path(*args):
    """Returns absolute path to the relative path provided"""
    return os.path.join(os.path.split(__file__)[0],*args)
    
def dir_from_path(path):
    """Returns directory path from file path"""
    return os.path.dirname(path)

def posix_path(*args):
    """Returns platform independent filename"""
    if platform.system() == 'Linux': 
        if len(args) > 1:
            return posixpath.join(*args)
        else:
            return args[0]
    elif platform.system() == 'Windows':
        if len(args) > 1:
            path = os.path.normpath(posixpath.join(*args))
        else:
            path = os.path.normpath(args[0])
        # remove any leading slash
        if path[0] == '\\':
            return path[1:]
        else:
            return path
        
def open_file(filename, abs=False):
    if abs:
        filename_mod = abs_path(filename)
    else:
        filename_mod = filename
        
    if platform.system() == 'Linux':
        ret_code = subprocess.call(('xdg-open', filename_mod))
        if ret_code:
            log.error('open_file - Error opening file - ' + filename_mod)
    elif platform.system() == 'Windows':
        os.startfile(filename_mod)
        
def get_file_path_from_dnd_dropped_uri(uri):
    # Get the path to file
    path = ""
    if uri.startswith('file:\\\\\\'): # windows
        path = uri[8:] # 8 is len('file:///')
    elif uri.startswith('file://'): # nautilus, rox
        path = uri[7:] # 7 is len('file://')
    elif uri.startswith('file:'): # xffm
        path = uri[5:] # 5 is len('file:')

    path = url2pathname(path) # escape special chars
    path = path.strip('\r\n\x00') # remove \r\n and NULL

    return path

def clean_markup(text):
    """Clear markup text of special characters"""
    for splchar, replspelchar in zip(['&', '<', '>', ], ['&amp;', '&lt;', '&gt;']):
        text = text.replace(splchar, replspelchar)
    return text

def font_str_parse(string):
    pango_font = Pango.FontDescription.from_string(string)
    family = pango_font.get_family().lower()
    size = pango_font.get_size()/Pango.SCALE
    return (family, size)

def font_str_encode(family, size):
    pango_font = Pango.FontDescription()
    pango_font.set_family(family)
    pango_font.set_family(size)
    string = pango_font.to_string()
    return string

def log_interpolate(points, num=10, prefix='point'):
    x = [x for x,y in points]
    y = [y for x,y in points]
    fit_func = interp1d(np.log10(x), np.log10(y), kind='quadratic')
    fit_func_mod = lambda x: list(10**(fit_func(np.log10(np.array(x)))))
    X = np.geomspace(min(x), max(x), num=num)
    Y = fit_func_mod(X)
    if prefix:
        points_int = [(prefix,x,y) for x,y in zip(X,Y)]
    else:
        points_int = [(x,y) for x,y in zip(X,Y)]
    return points_int

def log_interpolate_piecewise(points):
    
    def log_interpolate(x, y, num):
        fit_func = interp1d(np.log10(x), np.log10(y))
        fit_func_mod = lambda x: list(10**(fit_func(np.log10(np.array(x)))))
        X = np.geomspace(min(x), max(x), num=num)
        Y = fit_func_mod(X)
        return tuple(X), tuple(Y)

    I0, t0 = points[0]
    curve_i = [I0]
    curve_t = [t0]
    for I,t in points[1:]:
        if I0 != 0 and t0 != 0 and I !=0 and t != 0:
            deli = I/I0
            delt = t/t0
            dist_norm = max(deli, delt)
            if dist_norm > 1.5:
                num = int(20*np.log10(dist_norm))
                i_vals = [I0, I]
                t_vals = [t0, t]
                i_array_int, t_array_int = log_interpolate(i_vals, t_vals, num)
                curve_i += i_array_int[1:]
                curve_t += t_array_int[1:]
            else:
                curve_i += [I]
                curve_t += [t]
        else:
            curve_i += [I]
            curve_t += [t]
        I0, t0 = I, t
    points_int = tuple([(x,y) for x, y in zip(curve_i, curve_t)])
    return points_int

def open_library(filename):
    filename_abs = posix_path(USER_LIBRARY_DIR, filename)
    filename_abs_lib = abs_path('database', filename)
    if os.path.isfile(filename_abs):
        return filename_abs
    elif os.path.isfile(filename_abs_lib):
        return filename_abs_lib
    else:
        return None

# Field handling functions

def get_field_dict(field_type, caption, unit, value, 
                        max_chars=None, 
                        validation_func=None, 
                        selection_list=None, 
                        selection_image_list=None, 
                        decimal=6,
                        status_enable=True,
                        status_inactivate=True,
                        status_floating=False,
                        click_to_edit_message=None,
                        alter_structure=False,
                        alter_values_dict={},
                        tooltip=''):
    """
    ARGUMENTS:
    field_type:             Any of ('int','float','str','multiline_str','bool', 'font', 'graph')
    validation_func:        Function to validate input, should return validated result
    selection_list:         List of values to select from
    selection_image_list:   Corresponding images to load
    status_enable:          Whether to display field ?
    status_inactivate:      Whether to deactivate editing ?
    status_floating:        Whether to display field while floating (only used for element addition)
    click_to_edit_message:  Show a message before editing field
    alter_structure:        Whether field alters field structure of base element ?
    """
    field_dict = dict()
    field_dict['type'] = field_type
    field_dict['caption'] = caption
    field_dict['unit'] = unit
    field_dict['value'] = value
    field_dict['max_chars'] = max_chars
    field_dict['validation_func'] = validation_func
    field_dict['selection_list'] = selection_list
    field_dict['selection_image_list'] = selection_image_list
    field_dict['decimal'] = decimal
    field_dict['status_enable'] = status_enable
    field_dict['status_inactivate'] = status_inactivate
    field_dict['status_floating'] = status_floating
    field_dict['click_to_edit_message'] = click_to_edit_message
    field_dict['alter_structure'] = alter_structure
    field_dict['alter_values_dict'] = alter_values_dict
    field_dict['tooltip'] = tooltip
    return field_dict

def get_fields_trunc(fields):
    updated_fields = dict()
    for key, field in fields.items():
        updated_field = {'value': field['value'],
                         'type' : field['type']}
        updated_fields[key] = updated_field
    return updated_fields

def update_fields(reffields, newfields):
    updated = copy.deepcopy(reffields)
    for key in newfields:
        if key in reffields and updated[key]['type'] == newfields[key]['type']:
            updated[key]['value'] = newfields[key]['value']
    return updated
    
def get_fields_dict_trunc(fields_dict):
    updated_fields_dict = dict()
    for key, fields in fields_dict.items():
        updated_fields_dict[key] = get_fields_trunc(fields)
    return updated_fields_dict

def update_fields_dict(reffields_dict, newfields_dict):
    updated = copy.deepcopy(reffields_dict)
    for key in newfields_dict:
        if key in reffields_dict:
            updated[key] = update_fields(reffields_dict[key], newfields_dict[key])
    return updated

def set_field_selection_list(field, code, selection_list, modified={}):
    field[code]['selection_list'] = selection_list
    if (selection_list is not None) and (field[code]['value'] not in selection_list):
        modified[code] = field[code]['value'] 
        field[code]['value'] = selection_list[0]

def get_fields_from_params(parameters, modify_code=''):
    """Return feilds from paramater dict
        Order of parameters: caption, unit, value, [selection_list, tooltip, data_type, status_enable, default_values_dict]
            where
            default_values_dict: value -> default_values
                where
                default_values: parameter_key -> default value of parameter [if None, parameter is disabled]
    """
    fields = dict()
    default_parameter_dict = dict()  # default_values_dict combined for all active parameters
    
    def get_validation_func(decimal, min, max):
        def validation_func(text):
            try:
                validated = round(float(eval(text)), decimal)
                if min is not None:
                    if validated < min:
                        validated = min
                if max is not None:
                    if validated > max:
                        validated = max
            except:
                if min is not None:
                    validated = min
                else:
                    validated = 0
            return validated
        return validation_func

    for key, values in parameters.items():
        selection_list = None
        tooltip = ''
        data_type = 'float'
        validation_func = None
        status_enable = True
        alter_structure = False
        alter_values_dict = dict()
        caption, unit, value = values[0:3]
        if len(values) >= 4:
            selection_list = values[3]
        if len(values) >= 5:
            tooltip = values[4]
        if len(values) >= 6:
            # If type specified, use that
            if isinstance(values[5], str):
                data_type = values[5]
            # Else get additional parameters for float
            else:
                decimal, min, max = values[5]
                validation_func = get_validation_func(decimal, min, max)
        if len(values) >= 7:
            status_enable = values[6]
        if len(values) >= 8:
            alter_values_dict = values[7]
            alter_structure = True
        fields[modify_code+key] = get_field_dict(data_type, caption, unit, value, 
                                                selection_list=selection_list,
                                                tooltip=tooltip,
                                                status_inactivate=False,
                                                status_enable=status_enable,
                                                alter_structure=alter_structure,
                                                alter_values_dict=alter_values_dict,
                                                validation_func=validation_func)
    return fields

def update_params_from_params(parameters, new_parameters):
    for key, values in new_parameters.items():
        if key in parameters:
            parameters[key][2] = values[2]
            if len(parameters[key]) >= 7:
                parameters[key][6] = values[6]

def update_params_from_fields(parameters, fields):
    for key, field in fields.items():
        if key in parameters:
            parameters[key][2] = field['value']
            if len(parameters[key]) >= 7:
                parameters[key][6] = field['status_enable']

def validate_protection_data_struct(data_struct, code):
    check_keys_1 = {'type', 'parameters', 'data', 'graph_model'}
    if code in ['pcurve_l', 'pcurve_g']:
        prot_type = 'protection'
    elif code in ['dcurve']:
        prot_type = 'damage'
    else:
        return False
    if isinstance(data_struct, dict):
        keys = set(data_struct.keys())
        if check_keys_1 & keys == check_keys_1:
            if (isinstance(data_struct['type'], str) and data_struct['type'] == prot_type) and \
               (isinstance(data_struct['parameters'], dict)) and \
               (isinstance(data_struct['graph_model'], list)):
                return True
    return False

def get_blank_data_struct(data_type='protection'):
    data_struct = { 'type'          : data_type,
                    'parameters'    : {},
                    'data'          : {'curve_u': [], 'curve_l': []},
                    'graph_model'   : []}
    return data_struct

    
ELEMENT_FIELD = 0
ELEMENT_RESULT = 1
ELEMENT_PLACEHOLDER = 2

def elements_to_table(elements, col_codes, col_captions, code_sources, table_class=None, 
                      show_slno=True, show_element_class=True, modifyfunc=None, sum_cols=None):
    table = dict()
    clean = lambda x: clean_markup(str(x)).replace('\n','</br>')
    # Add unit line
    if show_slno:
        table['Sl.No.'] = ['']
    for col_caption in col_captions:
        table[col_caption] = ['']
    if show_element_class:
        table['Item Class'] = ['']
    # Add elements
    index = 1
    for element in elements:
        if show_slno:
            table['Sl.No.'].append(index)
        if show_element_class:
            table['Item Class'].append(element.name)
        for col_code, col_caption, code_source in zip(col_codes, col_captions, code_sources):
            # Select table
            if code_source == ELEMENT_FIELD:
                fields = element.fields
            elif code_source == ELEMENT_RESULT:
                fields = element.res_fields
            else:
                table[col_caption].append('')
                continue
            # Set value
            if col_code in fields:
                field = fields[col_code]
                table[col_caption][0] = field['unit']
                if field['type'] == 'graph':
                    if field['selection_list']:
                        table[col_caption].append(field['selection_list'][field['value']][0])
                    else:
                        table[col_caption].append(field['value'][0].replace('\n','</br>'))
                elif field['type'] == 'float':
                    table[col_caption].append(str(round(field['value'], field['decimal'])))
                else:
                    table[col_caption].append(clean(field['value']))
            else:
                table[col_caption].append('')
        index += 1
    if modifyfunc:
        modifyfunc(table)
    if sum_cols:
        for index in sum_cols:
            # Get sum
            def float_conv(value):
                try:  # try evaluating float
                    return float(value)
                except:
                    return 0
                
            sum_of_col = sum(map(float_conv, table[col_captions[index]][1:])) # Ignore unit row
            # Add blank row
            if show_slno:
                table['Sl.No.'].append('&Sigma;')
            for col_caption in col_captions:
                table[col_caption].append('')
            if show_element_class:
                table['Item Class'].append('')
            table[col_captions[index]][-1] = sum_of_col
            
    return pd.DataFrame(table).to_html(index=False, escape=False, classes=table_class)
    
def fields_to_table(fields, insert_image=True, insert_graph=True):
    table = {'Sl.No.':[], 'Description': [], 'Value': [], 'Unit': []}
    clean = lambda x: clean_markup(str(x)).replace('\n','</br>')
    index = 1
    for field in fields.values():
        if field['status_enable'] == True:
            if field['type'] not in ('graph', 'data', 'heading'):
                table['Sl.No.'].append(index)
                table['Description'].append(clean(field['caption']))
                table['Unit'].append(field['unit'])
                if field['type'] == 'float':
                    value = str(round(field['value'], field['decimal']))
                else:
                    value = clean(field['value'])
                if field['selection_list'] and field['selection_image_list'] and insert_image:
                    sel_index = field['selection_list'].index(field['value'])
                    image_path = field['selection_image_list'][sel_index]
                    if image_path:
                        image_file = abs_path('icons', image_path)
                        data_uri = base64.b64encode(open(image_file, 'rb').read()).decode('utf-8')
                        img_tag = "<div class='parent flex-parent'><div class='child flex-left'><img class='img_table' src='data:image/svg+xml;base64,{0}'></div><div class='child flex-right'>{1}</div></div>".format(data_uri, value)
                        table['Value'].append(img_tag)
                    else:
                        table['Value'].append(value)
                else:
                    table['Value'].append(value)
                index += 1
            elif field['type'] == 'heading':
                table['Sl.No.'].append('')
                table['Description'].append('<b>' + clean(field['caption']) + '</b>')
                table['Unit'].append(field['unit'])
                table['Value'].append(clean(field['value']))
            elif field['type'] == 'data':
                table['Sl.No.'].append(index)
                table['Description'].append(clean(field['caption']))
                table['Unit'].append(field['unit'])
                img_tag = ''
                graph_fields = ''
                # Add graphimage
                if insert_graph:
                    from .view.graph import GraphImage
                    xlim, ylim, xlabel, ylabel, graph_params = field['graph_options']
                    title = field['value']['graph_model'][0]
                    graph_models = field['value']['graph_model'][1]
                    graph_image = GraphImage(xlim, ylim, title, xlabel, ylabel, graph_params)
                    graph_image.add_plots(graph_models)
                    img_tag = graph_image.get_embedded_html_image(figsize=(200, 200))
                # Add graph options
                param_fields = get_fields_from_params(field['value']['parameters'])
                for param_key, param_field in param_fields.items():
                    caption = param_field['caption']
                    unit = param_field['unit']
                    value = param_field['value']
                    status_enable = param_field['status_enable']
                    param_type = param_field['type']
                    if status_enable:
                        if param_type == 'heading':
                            graph_fields += '</br><b>' + clean(caption) + '</b>'
                        else:
                            graph_fields += '</br>' + clean(caption + ' : ' + str(value) + ' ' + unit)
                table['Value'].append(img_tag + graph_fields.lstrip('</br>'))
                index += 1
            else:
                table['Sl.No.'].append(index)
                table['Description'].append(clean(field['caption']))
                table['Unit'].append(field['unit'])
                if field['selection_list']:
                    table['Value'].append(field['selection_list'][field['value']][0])
                else:
                    title = field['value'][0]
                    if insert_graph:
                        from .view.graph import GraphImage
                        xlim, ylim, xlabel, ylabel, graph_params = field['graph_options']
                        graph_models = field['value'][1]
                        graph_image = GraphImage(xlim, ylim, title, xlabel, ylabel, graph_params)
                        graph_image.add_plots(graph_models)
                        img_tag = graph_image.get_embedded_html_image(figsize=(200, 200))
                        table['Value'].append(img_tag)
                    else:
                        table['Value'].append(clean(title))
                index += 1
    return pd.DataFrame(table).to_html(index=False, escape=False, classes='element_fields')
    
def params_to_table(fields_list, titles):
    """Return pandas table with comparison of field values"""

    def get_frame(fields, caption):
        table = {'REF':[], 'DESCRIPTION': [], 'UNIT': [], caption: []}
        for ref, field in fields.items():
            if field['status_enable'] == True:
                table['REF'].append(ref)
                table['DESCRIPTION'].append(field['caption'])
                table['UNIT'].append(field['unit'])
                if field['type'] == 'float':
                    value = str(round(field['value'], field['decimal']))
                else:
                    value = field['value']
                table[caption].append(value)
        return pd.DataFrame(table)
    
    table_combined = None
    for fields, title in zip(fields_list, titles):
        table = get_frame(fields, title)
        if table_combined is not None:
            table_combined = pd.merge(table_combined, table, how='outer')
        else:
            table_combined = table
    columns = ['REF', 'DESCRIPTION', 'UNIT', *titles]
    return table_combined[columns]
    
def get_uid():
    """Get unique id as identifier"""
    return str(uuid())

# Default settings

default_program_settings = {'Defaults':{'drawing_field_dept':    get_field_dict('str', 'Responsible department', '', '', status_inactivate=False),
                            'drawing_field_techref': get_field_dict('str', 'Technical reference', '', '', status_inactivate=False),
                            'drawing_field_created': get_field_dict('str', 'Created by', '', '', status_inactivate=False),
                            'drawing_field_approved':get_field_dict('str', 'Approved by', '', '', status_inactivate=False),
                            'drawing_field_lang':    get_field_dict('str', 'Language code', '', 'en', status_inactivate=False),
                            'drawing_field_address': get_field_dict('multiline_str', 'Address', '', 'WING\nORGANISATION\nLOCATION', status_inactivate=False)},
                            'Interface':{'dark_mode' : get_field_dict('bool', 'Enable dark theme', '(Requires program restart to take effect)', False, status_inactivate=False),
                                         'advanced_mode' : get_field_dict('bool', 'Enable advanced mode', '(Requires program restart to take effect)', False, status_inactivate=False),
                                         'show_graphs' : get_field_dict('bool', 'Show graphs in field view', '(Requires program restart to take effect)', True, status_inactivate=False),
                                         'drawing_font':    get_field_dict('font', 'Drawing Font', '', SCHEM_FONT_FACE + ' ' + str(SCHEM_FONT_SIZE), status_inactivate=False),
                                         'graph_font':    get_field_dict('font', 'Graph Font', '', GRAPH_FONT_FACE + ' ' + str(GRAPH_FONT_SIZE), status_inactivate=False),
                                         'report_font':    get_field_dict('font', 'Report Font', '', REPORT_FONT_FACE + ' ' + str(REPORT_FONT_SIZE), status_inactivate=False),
                                         'report_graph_font':    get_field_dict('font', 'Report Graph Font', '', REPORT_GRAPH_FONT_FACE + ' ' + str(REPORT_GRAPH_FONT_SIZE), status_inactivate=False)},
                            'Paths': {'library_path':    get_field_dict('path', 'User Library', '', '')}}
                                         
default_project_settings = {'Information': {'project_name': get_field_dict('str', 'Project Name', '', 'PROJECT', status_inactivate=False),
                            'drawing_field_dept':    get_field_dict('str', 'Responsible department', '', '', status_inactivate=False),
                            'drawing_field_techref': get_field_dict('str', 'Technical reference', '', '', status_inactivate=False),
                            'drawing_field_created': get_field_dict('str', 'Created by', '', '', status_inactivate=False),
                            'drawing_field_approved':get_field_dict('str', 'Approved by', '', '', status_inactivate=False),
                            'drawing_field_lang':    get_field_dict('str', 'Language code', '', 'en', status_inactivate=False),
                            'drawing_field_address': get_field_dict('multiline_str', 'Address', '', 'WING\nORGANISATION\nLOCATION', status_inactivate=False)},
                             
                             'Simulation':{'run_diagnostics' : get_field_dict('bool', 'Run diagnostics', '', True, status_inactivate=False),
                             # Power flow
                             'sub_head_power_flow' : get_field_dict('heading', 'Power Flow', '', '', status_inactivate=False),
                             'run_powerflow' : get_field_dict('bool', 'Run time series power flow', '', False, status_inactivate=False),
                             'pf_method' : get_field_dict('str', 'Power flow method', '', 'Power flow', 
                                                          status_inactivate=False,
                                                          selection_list=['Power flow', 'Power flow with diversity', 'Time series'] ),
                             'power_flow_3ph' : get_field_dict('bool', 'Enable assymetric power flow calculation', '', False, status_inactivate=False),
                             # Short circuit
                             'sub_head_sc' : get_field_dict('heading', 'Short Circuit Calculations', '', '', status_inactivate=False),
                             'run_sc_sym' : get_field_dict('bool', 'Run symmetric short circuit calculation', '', False, status_inactivate=False),
                             'run_sc_gf' : get_field_dict('bool', 'Run line to ground short circuit calculation', '', False, status_inactivate=False),
                             'show_impedances' : get_field_dict('bool', 'Display short circuit impedance values', '', False, status_inactivate=False),
                             # Simulation parameters
                             'export_results' : get_field_dict('bool', 'Export results of simulation', '', False, status_enable=False),
                             'sub_head_sim_param' : get_field_dict('heading', 'Simulation Parameters', '', '', status_inactivate=False),
                             'lv_tol_percent': get_field_dict('float', 'Grid voltage tolerance', '%', 6, selection_list=[6,10], status_inactivate=False),
                             'grid_frequency': get_field_dict('float', 'Grid Frequency', 'Hz', 50, selection_list=[50,60], status_inactivate=False),
                             'r_fault_ohm': get_field_dict('float', 'Fault resistance', 'Ohm', 0, status_inactivate=False),
                             'x_fault_ohm': get_field_dict('float', 'Fault reactance', 'Ohm', 0, status_inactivate=False)},

                             'Reports':{'export_elements' : get_field_dict('bool', 'Export elements', '', True, status_inactivate=False),
                                        'export_boq' : get_field_dict('bool', 'Export BOQ', '', True, status_inactivate=False),
                                        'export_loadprofiles' : get_field_dict('bool', 'Export load profiles', '', True, status_inactivate=False),
                                        'export_analysis' : get_field_dict('bool', 'Export analysis results', '', True, status_inactivate=False),
                                        'export_graphs' : get_field_dict('bool', 'Include graphs in report', '', False, status_inactivate=False),},
                             
                             'Rules Check':{'line_max_loss' : get_field_dict('float', 'Maximum line loss', '%', 3, status_inactivate=False),
                                            'max_disc_time' : get_field_dict('float', 'Maximum disconnection time for faults', 's', 5, status_inactivate=False),
                                            'max_voltage_drop' : get_field_dict('float', 'Maximum voltage drop at loads', '%', 5, status_inactivate=False)}}

loadprofile_blank_fields = {'name': get_field_dict('str', 'Title', '', ''),
                            'name1': get_field_dict('str', 'Sub Title', '', '')}
for slno in range(0,24):
    loadprofile_blank_fields[str(slno)] = get_field_dict('float', str(slno), '', 0)

# Cairo drawing functions

def rgb2hex(r,g,b,a=None):
    if a:
        return "#{:02x}{:02x}{:02x}{:02x}".format(int(r*255),int(g*255),int(b*255),int(a*255))
    else:
        return "#{:02x}{:02x}{:02x}".format(int(r*255),int(g*255),int(b*255))

def hex2rgb(hexcode):
    if len(hexcode) == 7:
        r = int(hexcode[1:3],16)/255 
        g = int(hexcode[3:5],16)/255
        b = int(hexcode[5:7],16)/255
        a = 1
    elif len(hexcode) == 9:
        r = int(hexcode[1:3],16)/255 
        g = int(hexcode[3:5],16)/255
        b = int(hexcode[5:7],16)/255
        a = int(hexcode[7:9],16)/255
    return (r, g, b, a)
    
def rect_from_points(*points):
    if len(points) > 1:
        maxx = minx = points[0][0]
        maxy = miny = points[0][1]
        for point in points[1:]:
            x, y = point
            maxx = x if x > maxx else maxx
            maxy = y if y > maxy else maxy
            minx = x if x < minx else minx
            miny = y if y < miny else miny
        x0 = int(minx)
        y0 = int(miny)
        width = int(maxx-minx)
        height = int(maxy-miny)
        return (x0, y0, width, height)
    else:
        return None
    
def draw_rectangle(context, x, y, w, h, radius=0, color=COLOR_NORMAL, line_style='solid', dash_pattern=[10,10], stroke_width=STROKE_WIDTH_NORMAL, fill=False):
    (r,g,b,a) = hex2rgb(color)
    context.save()
    context.set_source_rgba(r, g, b, a)
    context.set_line_width(stroke_width)
    if line_style == 'solid':
        context.arc(x+radius, y+radius, radius, math.pi, 3*math.pi/2)
        context.arc(x+w-radius, y+radius, radius, 3*math.pi/2, 0)
        context.arc(x+w-radius, y+h-radius, radius, 0, math.pi/2)
        context.arc(x+radius, y+h-radius, radius, math.pi/2, math.pi)
        context.close_path()
        if fill:
            context.fill()
        else:
            context.stroke()
    elif line_style == 'dashed':
        context.set_dash(dash_pattern)
        context.arc(x+radius, y+radius, radius, math.pi, 3*math.pi/2)
        context.arc(x+w-radius, y+radius, radius, 3*math.pi/2, 0)
        context.arc(x+w-radius, y+h-radius, radius, 0, math.pi/2)
        context.arc(x+radius, y+h-radius, radius, math.pi/2, math.pi)
        context.close_path()
        context.stroke()
    context.restore()
    
def draw_line(context, x0, y0, x1, y1, color=COLOR_NORMAL, dash_pattern=[], stroke_width=STROKE_WIDTH_NORMAL):
    (r,g,b,a) = hex2rgb(color)
    context.save()
    context.set_source_rgba(r, g, b, a)
    context.set_line_width(stroke_width)
    context.set_dash(dash_pattern)
    context.move_to(x0,y0)
    context.line_to(x1,y1)
    context.stroke()
    context.restore()

def draw_text(context, text, x, y, color=COLOR_NORMAL, fontname='Sans', fontsize=12, fontweight=SCHEM_FONT_WEIGHT, alignment='left'):
    """Draw text using Pango"""
    context.save()
    context.translate(x,y)
    (r,g,b,a) = hex2rgb(color)
    context.set_source_rgba(r, g, b, a)
    layout = PangoCairo.create_layout(context)
    font = Pango.FontDescription(fontname + " " + str(fontsize))
    font.set_weight(fontweight)
    layout.set_font_description(font)
    layout.set_text(text)
    # Modify coordinates for alignment
    (width, height) = layout.get_size()
    w = Pango.units_to_double(width)
    h = Pango.units_to_double(height)
    if alignment == 'center': 
        layout.set_alignment(Pango.Alignment.CENTER)
        dx = -w/2
    elif alignment == 'right':
        layout.set_alignment(Pango.Alignment.RIGHT)
        dx = -w
    else:
        layout.set_alignment(Pango.Alignment.LEFT)
        dx = 0
    context.translate(dx,0)
    PangoCairo.show_layout(context, layout)
    context.restore()

    return (x+dx, y, w, h)

def get_image_from_path(path, height=None, inverted=False, mode='image'):
            
    def draw_pixel(pixels, x, y, rgba, width, height):
        # make sure pixel data is reasonable
        x = min(x, width)
        y = min(y, height)
        r = min(rgba[0], 255)
        g = min(rgba[1], 255)
        b = min(rgba[2], 255)
        a = min(rgba[3], 255)
        # insert pixel data at right location in bytes array
        i = y*width + x
        pixels[i*4 + 0] = r
        pixels[i*4 + 1] = g
        pixels[i*4 + 2] = b
        pixels[i*4 + 3] = a
    if height:
        img = GdkPixbuf.Pixbuf.new_from_file_at_scale(path, -1, 400,True)
    else:
        img = GdkPixbuf.Pixbuf.new_from_file(path)
    if inverted:
        pixels = bytearray(img.get_pixels())
        width = img.props.width
        height = img.props.height
        for row_num, row in enumerate(zip(*(iter(img.get_pixels()),) *img.get_rowstride())):
            for col_num, pixel in enumerate(zip(*(iter(row),) * 4)):
                r, g, b, a = pixel
                draw_pixel(pixels, col_num, row_num, [255-r, 255-g, 255-b, a], width, height)
        img_new = GdkPixbuf.Pixbuf.new_from_bytes(GLib.Bytes(pixels), GdkPixbuf.Colorspace.RGB, True, 8,
                                                width, height, img.get_rowstride())
    else:
        img_new = img
    if mode == 'pixbuf':
        return img_new
    elif mode == 'image':
        icon = Gtk.Image()
        icon.set_from_pixbuf(img_new)
        return icon

